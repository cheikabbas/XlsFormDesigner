import os

import wx
import openpyxl


class Export(wx.Panel):
    """Onglet d'exportation"""

    def __init__(self, parent, survey, choices, settings):
        super(Export, self).__init__(parent)

        self.survey = survey
        self.choices = choices
        self.settings = settings

        vbox = wx.BoxSizer(wx.VERTICAL)
        hbox = wx.BoxSizer(wx.HORIZONTAL)

        self.ctrl_dir = wx.TextCtrl(self, style=wx.TE_READONLY, value="Aucun repertoire d'exportation choisi")

        self.btn_select_dir = wx.Button(self, label="Choisir un répertoire")
        self.btn_select_dir.Bind(wx.EVT_BUTTON, self.on_select_directory)

        hbox.Add(self.ctrl_dir, proportion=1, flag=wx.EXPAND | wx.ALL, border=5)
        hbox.Add(self.btn_select_dir, flag=wx.ALL, border=5)

        label_file = wx.StaticText(self, label="Nom du fichier")
        self.ctrl_file = wx.TextCtrl(self, value=".xlsx")

        vbox.Add(hbox,  0, wx.ALL | wx.EXPAND, 5)
        vbox.Add(label_file,  0, wx.ALL | wx.EXPAND, 5)
        vbox.Add(self.ctrl_file, 0, wx.ALL | wx.EXPAND, 5)

        self.export_btn = wx.Button(self, label="Exporter en Excel")
        self.export_btn.SetBackgroundColour("#05f762")
        self.export_btn.Bind(wx.EVT_BUTTON, self.on_export)

        vbox.Add(self.export_btn, 0, wx.ALL, 5)

        self.SetSizer(vbox)

    def on_select_directory(self, event):
        """Ouvre un dialogue de sélection de dossier"""
        with wx.DirDialog(self, "Sélectionnez un répertoire", style=wx.DD_DEFAULT_STYLE) as dialog:
            if dialog.ShowModal() == wx.ID_OK:
                self.ctrl_dir.SetValue(dialog.GetPath())

    def on_export(self, event):
        """Exporter le contenu du grid vers un fichier Excel."""
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        # Créer les feuilles
        survey = wb.create_sheet("survey")
        choices = wb.create_sheet("choices")
        settings = wb.create_sheet("settings")

        # Headers
        for data, sheet in zip([self.survey, self.choices, self.settings], [survey, choices, settings]):
            for col in range(data.grid.GetNumberCols()):
                header = data.grid.GetColLabelValue(col)
                sheet.cell(row=1, column=col+1, value=header)

        for row in range(self.survey.grid.GetNumberRows()):
            for col in range(self.survey.grid.GetNumberCols()):
                value = self.survey.grid.GetCellValue(row, col)
                survey.cell(row=row + 2, column=col + 1, value=value)

        for row in range(self.choices.grid.GetNumberRows()):
            for col in range(self.choices.grid.GetNumberCols()):
                value = self.choices.grid.GetCellValue(row, col)
                choices.cell(row=row + 2, column=col + 1, value=value)

        for row in range(self.settings.grid.GetNumberRows()):
            for col in range(self.settings.grid.GetNumberCols()):
                value = self.settings.grid.GetCellValue(row, col)
                settings.cell(row=row + 2, column=col + 1, value=value)

        # Sauvegarder le fichier Excel
        filename = os.path.join(self.ctrl_dir.GetValue(), self.ctrl_file.GetValue())
        wb.save(filename)
        wx.MessageBox(f"Le fichier {filename} a été créé avec succès !", "Export réussi", wx.OK | wx.ICON_INFORMATION)
